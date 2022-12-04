import pyautogui as abrirsite
from openpyxl import load_workbook
from tkinter import Button, filedialog
import tkinter as tk

# pyinstaller --onefile --console .\docusign4tk.py


print('\n\033[1;32mSoftware by Anderson Marinho \033[m')
print('\033[1;32mVersão 1.0 \033[m')
print('\n')

print("|   --> Robot ENCONTROU esses dados para fazer ação <--                   ")
print('|   --> Siga as instruções <--                   ')
print('\n\033[1;33mSelecione o arquivo de base de dados na tela tk.\033[m')
#11

root= tk.Tk()
canvas1 = tk.Canvas(root, width = 300, height = 300, bg = 'lightsteelblue')
canvas1.pack()
abrirsite.sleep(1)
with abrirsite.hold('alt'):
    abrirsite.press('tab', presses=2)

def getExcel ():
    global nome_arquivo
    import_file_path = filedialog.askopenfilename()
    nome_arquivo = import_file_path
    planilha_aberta = nome_arquivo
    print(nome_arquivo)
    browseButton_Excel.destroy()
    root.destroy()
    abrirsite.sleep(1)    


    # nome_arquivo = input(r'Digite Caminho da BaseDados: ') + str('\\BaseSign.xlsx')
    # nome_arquivo = r'N:\CromexOneDrive\Cromex S A\Administração de Pessoal - ADMINISTRAÇÃO DE PESSOAL E BENEFÍCIOS\FISCALIZAÇÃO\_LGPD\_ENVIO\BaseSign.xlsx'
    # print(nome_arquivo)
    


    planilha_aberta = load_workbook(filename=nome_arquivo)
    sheet_selecionada = planilha_aberta[planilha_aberta.sheetnames[0]]
    # coluna1 = input("\n\033[1;34mDigite a letra da coluna com [NOME]   ex. 'A': \033[m")
    while True:
        try:
            coluna1 = input("\n\033[1;34mDigite a letra da coluna onde está o NOME   ex:|'A'| \033[m")
            if coluna1 == 'A':
                break
        except:
            print("\n\033[1;31mERRO: Digite uma letra válida!\033[m")
            return coluna1

    # coluna2 = input("\n\033[1;34mDigite a letra da coluna com [E-MAIL] ex. 'Q': \033[m")
    while True:
        try:
            coluna2 = input("\n\033[1;34mDigite a letra da coluna onde está o E-MAIL ex:|'B' ou 'Q'| \033[m")
            if coluna2 == 'Q':
                break
            elif coluna2 == 'B':
                break
        except:
            print("\n\033[1;31mERRO: Digite uma letra válida!\033[m")
            return coluna2


    while True:
        linhaWS = input("\nA partir da linha tal: ")
        qtdLinhas = input("\nQuantas linhas você quer: ")

        abrirsite.sleep(1)
        abrirsite.hotkey('alt', 'tab')

        abrirsite.sleep(1)
        input('\n\033[1;32mPosicione o MOUSE no campo do site e tecle ENTER para continuar... \033[m')
        abrirsite.sleep(1)
        abrirsite.hotkey('alt', 'tab')
        abrirsite.sleep(1.5)
        


        contador = 0
        for linha in range(int(linhaWS), len(sheet_selecionada[coluna1]) + 1):
            nomeCompleto = sheet_selecionada[coluna1+'%s' % linha].value
            emailNome = sheet_selecionada[coluna2+'%s' % linha].value 
            
            nomeCompleto = str(nomeCompleto)
            emailNome = str(emailNome)
            qtdLinhas = int(qtdLinhas)
            

            abrirsite.write(nomeCompleto)
            abrirsite.sleep(1)
            abrirsite.press('tab', presses=2)
            abrirsite.sleep(1)
            abrirsite.write(emailNome)
            abrirsite.sleep(1)
            abrirsite.press('tab', presses=7)
            abrirsite.sleep(1)
            abrirsite.press('enter')
            abrirsite.sleep(1)
            contador = contador + 1
            contador
            print(nomeCompleto, ' - ', emailNome)
            if contador == int(qtdLinhas):
                print(f'Foram digitado(s)', qtdLinhas, 'funcionári(os)')
                break
        sair = int(input('\n\033[1;33mContinuar no RPA? [1]sim ou [2]não: \033[m'))
        if sair == 1:
            linhaWS
            continue
        else:
            break

    print('\n\33[1;31mFim da execução... \n\033[m')
browseButton_Excel = tk.Button(text='Import BaseDados BaseSign Excel', command=getExcel, bg='green', fg='white', font=('helvetica', 10, 'bold'))
canvas1.create_window(150, 150, window=browseButton_Excel)

root.mainloop()
